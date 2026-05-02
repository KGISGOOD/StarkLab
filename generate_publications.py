import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 資料 ──────────────────────────────────────────────────────────────────────

journal_papers = [
    {
        "type": "期刊論文",
        "title": "Temperature-dependent electroluminescence efficiency in blue InGaN–GaN light-emitting diodes with different well widths",
        "authors": "CH Wang, JR Chen, CH Chiu, HC Kuo, YL Li, TC Lu, SC Wang",
        "year": 2010,
        "venue": "IEEE Photonics Technology Letters",
        "volume_issue": "22 (4)",
        "pages": "236-238",
        "conference": "",
        "location": "",
    },
    {
        "type": "期刊論文",
        "title": "Study of the optical effects of nanostructure embedded GaN light emitting diodes formed by nanorod template overgrowth",
        "authors": "CY Wang, YC Lee, PS Lee, CH Chiu, CH Kuo, ML Wu",
        "year": 2013,
        "venue": "Thin Solid Films",
        "volume_issue": "539",
        "pages": "245-250",
        "conference": "",
        "location": "",
    },
    {
        "type": "期刊論文",
        "title": "Long text to image converter for financial reports",
        "authors": "CH Chiu, YC Tsai, HL Chen",
        "year": 2021,
        "venue": "International Journal of Data Mining, Modelling and Management",
        "volume_issue": "13 (3)",
        "pages": "211-230",
        "conference": "",
        "location": "",
    },
    {
        "type": "期刊論文",
        "title": "Rolling Machine Learning and Multi-Method Feature Selection: Portfolio Evidence from Taiwan's Monthly Revenue Disclosure System",
        "authors": "Chia-Hao Chiu* (邱嘉豪)",
        "year": 2026,
        "venue": "Taiwan Journal of Applied Economics (臺灣應用經濟學刊)",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
]

conference_papers = [
    {
        "type": "研討會論文",
        "title": "Enhancing the brightness of GaN light-emitting diodes by manipulating the illumination direction in the photoelectrochemical process",
        "authors": "HP Shiao, CY Wang, ML Wu, CH Chiu",
        "year": 2010,
        "venue": "IEEE Photonics Technology Letters",
        "volume_issue": "22 (22)",
        "pages": "1653-1655",
        "conference": "",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "The Improvement of Efficiency and Uniformity in Non-image LED Illumination for Field-Sequential-Color Pico-projector",
        "authors": "SD Jiang, CH Chiu, HC Wu, PS Lee, YC Lee, WS Sun, ML Wu",
        "year": 2011,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "The modulation of LEDs driving current and duration ratio in application of Color-Sequential Pico-Projector",
        "authors": "CH Chiu, SD Jiang, PS Lee, YC Lee, WS Sun, ML Wu",
        "year": 2011,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "Color-Temperature Modulation of RGB LEDs and Its Application of Color-Sequential Pico-Projectors",
        "authors": "邱嘉豪",
        "year": 2011,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "使用賽局理論分析挑選團體成員之戰略達到高滿意度配對之研究",
        "authors": "",
        "year": 2017,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "GCCCE2017",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "Use Text Mining for Financial Reports Analysis: Long Text to Image Converter",
        "authors": "CH Chiu, YC Tsai",
        "year": 2020,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "2020 the 6th International Conference on Communication and Information Processing",
        "location": "Tokyo",
    },
    {
        "type": "研討會論文",
        "title": "Predicting Period Stock Spread Ranking Using Revenue Indicators and Machine Learning Techniques",
        "authors": "CH Chiu, YC Tsai",
        "year": 2021,
        "venue": "IOP Conference Series: Earth and Environmental Science",
        "volume_issue": "704 (1)",
        "pages": "012014",
        "conference": "",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "Practical Impact of ChatGPT in Introduction to Computer Science Course: Exam Performance and Real Learning Outcomes",
        "authors": "Chia-En Wu, Chia-Hao Chiu",
        "year": 2024,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "GCCCE2024",
        "location": "Chongqing, China",
    },
    {
        "type": "研討會論文",
        "title": "Practical Impact of ChatGPT in Introduction to Computer Science Course: Exam Performance and Real Learning Outcomes",
        "authors": "",
        "year": 2024,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "ICIET",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "ChatGPT於計算機概論課程中的實際效果：測驗成效以及真實學習成效",
        "authors": "邱嘉豪",
        "year": 2024,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "2024全國科普論壇",
        "location": "高雄, 台灣",
    },
    {
        "type": "研討會論文",
        "title": "A Comprehensive Analysis of ChatGPT-Assisted Learning: A Systematic Exploration from Learning Outcomes to Student Behavior Patterns",
        "authors": "Chiao-Hsi Hsiao; Chia-Hao Chiu",
        "year": 2025,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "ICAIE",
        "location": "Suzhou, China",
    },
    {
        "type": "研討會論文",
        "title": "The Impact of Gamified Learning on Financial Literacy and Behavioral Change",
        "authors": "CH Chiu",
        "year": 2025,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "2025 5th International Conference on Artificial Intelligence and Education (ICAIE)",
        "location": "Suzhou, China",
    },
    {
        "type": "研討會論文",
        "title": "Practical Impact of ChatGPT in Introduction to Computer Science Course: Exam Score and Real Learning Effectiveness",
        "authors": "Chia-Hao Chiu",
        "year": 2025,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "IEEE-ICIET",
        "location": "Fukuyama, Japan",
    },
    {
        "type": "研討會論文",
        "title": "Integrating Large Language Models and Vector Databases for Podcast Semantic Search",
        "authors": "I-Ting Lin, Chia-Hao Chiu, Bo-Tang Liao, Bo-Jyun Chen",
        "year": 2025,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "2025 IEEE 14th International Workshop on Computational Intelligence and Applications",
        "location": "Hiroshima, Japan",
    },
    {
        "type": "研討會論文",
        "title": "AI與理學的跨域探索：一場關於際關係的實驗教育旅程",
        "authors": "邱嘉豪、Wei Tzu-Hsiang",
        "year": 2025,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "2025年第九屆實驗教育國際研討會",
        "location": "台灣、台北",
    },
    {
        "type": "研討會論文",
        "title": "生成式 AI 輔助下不同作業型態對大學計算機概論學習內化之影響：作業導向與報告導向策略的準實驗比較",
        "authors": "邱嘉豪*, 陳俊豪",
        "year": 2026,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "GCCCE 2026 (第30屆全球華人計算機教育應用大會)",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "Beyond the Veneer of Competence: How Structured Prompting Fosters Knowledge Transfer in LLM-Augmented Education",
        "authors": "I-Ting Lin, Chia-Hao Chiu*, Ko-I Lee",
        "year": 2026,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "IAIT 2026 (The 14th International Conference on Advances in Information Technology)",
        "location": "",
    },
    {
        "type": "研討會論文",
        "title": "Reinforcement Strategies in ChatGPT-Assisted Computational Thinking Education: A Comparison of Handwritten Practice and Oral Presentation",
        "authors": "Chia-Hao Chiu*",
        "year": 2026,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "APSCE TBICS 2026 (Festival of Learning / AI-CTE-STEM Track)",
        "location": "Kumamoto, Japan",
    },
    {
        "type": "研討會論文",
        "title": "Overcoming Traditional Podcast Search Limitations via RAG: An Investigation into Response Quality and Stability",
        "authors": "Kai Chun Cheng, Chia Hao Chiu",
        "year": 2026,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "APSCE TBICS 2026 (Festival of Learning)",
        "location": "Kumamoto, Japan",
    },
]

books = [
    {
        "type": "書籍",
        "title": "洞悉新興產業的23堂課－史塔克實驗室 StarkLab（上）",
        "authors": "邱嘉豪",
        "year": 2022,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
    {
        "type": "書籍",
        "title": "洞悉新興產業的23堂課－史塔克實驗室 StarkLab（下）",
        "authors": "邱嘉豪",
        "year": 2022,
        "venue": "",
        "volume_issue": "",
        "pages": "",
        "conference": "",
        "location": "",
    },
]

all_papers = journal_papers + conference_papers + books

# ── 樣式設定 ──────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # 深藍
JOURNAL_FILL  = PatternFill("solid", fgColor="D6E4F0")   # 淡藍
CONF_FILL     = PatternFill("solid", fgColor="E8F5E9")   # 淡綠
BOOK_FILL     = PatternFill("solid", fgColor="FFF8E1")   # 淡黃

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT   = Font(name="Calibri", size=10)

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

def apply_header(ws, headers, col_widths):
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[1].height = 30

def apply_row(ws, row_idx, values, fill):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 45

# ── 工作表 1：全部論文總表 ────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws_all = wb.active
ws_all.title = "全部論文總表"

headers_all = ["#", "類型", "年份", "論文 / 書名", "作者", "期刊 / 書系", "卷期", "頁碼", "研討會名稱", "地點"]
col_widths_all = [4, 10, 6, 55, 35, 35, 10, 10, 45, 20]
apply_header(ws_all, headers_all, col_widths_all)

TYPE_FILL = {
    "期刊論文": JOURNAL_FILL,
    "研討會論文": CONF_FILL,
    "書籍": BOOK_FILL,
}

for i, p in enumerate(all_papers, 1):
    fill = TYPE_FILL.get(p["type"], JOURNAL_FILL)
    apply_row(ws_all, i + 1, [
        i,
        p["type"],
        p["year"],
        p["title"],
        p["authors"],
        p["venue"],
        p["volume_issue"],
        p["pages"],
        p["conference"],
        p["location"],
    ], fill)

ws_all.freeze_panes = "A2"

# ── 工作表 2：期刊論文 ────────────────────────────────────────────────────────

ws_j = wb.create_sheet("期刊論文")
headers_j = ["#", "年份", "論文標題", "作者", "期刊名稱", "卷期", "頁碼"]
col_widths_j = [4, 6, 60, 40, 40, 10, 10]
apply_header(ws_j, headers_j, col_widths_j)

for i, p in enumerate(journal_papers, 1):
    apply_row(ws_j, i + 1, [
        i, p["year"], p["title"], p["authors"],
        p["venue"], p["volume_issue"], p["pages"],
    ], JOURNAL_FILL)

ws_j.freeze_panes = "A2"

# ── 工作表 3：研討會論文 ──────────────────────────────────────────────────────

ws_c = wb.create_sheet("研討會論文")
headers_c = ["#", "年份", "論文標題", "作者", "研討會名稱", "地點", "期刊 / 會刊", "卷期", "頁碼"]
col_widths_c = [4, 6, 55, 35, 48, 20, 35, 10, 10]
apply_header(ws_c, headers_c, col_widths_c)

for i, p in enumerate(conference_papers, 1):
    apply_row(ws_c, i + 1, [
        i, p["year"], p["title"], p["authors"],
        p["conference"], p["location"],
        p["venue"], p["volume_issue"], p["pages"],
    ], CONF_FILL)

ws_c.freeze_panes = "A2"

# ── 工作表 4：書籍 ────────────────────────────────────────────────────────────

ws_b = wb.create_sheet("書籍")
headers_b = ["#", "年份", "書名", "作者"]
col_widths_b = [4, 6, 60, 20]
apply_header(ws_b, headers_b, col_widths_b)

for i, p in enumerate(books, 1):
    apply_row(ws_b, i + 1, [
        i, p["year"], p["title"], p["authors"],
    ], BOOK_FILL)

ws_b.freeze_panes = "A2"

# ── 儲存 ──────────────────────────────────────────────────────────────────────

output_path = "C:/Users/Bandai/Desktop/StarkLab_Publications.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")
print(f"     Total {len(all_papers)} rows (Journal={len(journal_papers)}, Conference={len(conference_papers)}, Book={len(books)})")
